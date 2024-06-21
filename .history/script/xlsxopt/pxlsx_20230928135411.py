import os
import sys
import openpyxl

from script.utils.utils import utils
from script.utils.utilsword import utilsWord
from script.utils.utilsfile import utilsFile
from script.xlsxopt.basexlsx import bXlsx
from script.trans.pelbstrans import pTrans
from script.contentmgr import contentMgr
from script.xlsxopt.txtopt import tOpt

class PelbsXlsx:
    def __init__(self):
        self.unit_name = ""
    
    def create_audio_excel(self, res_texture_path):
        wb = openpyxl.Workbook()   # 创建excel文件
        temp_xls_file = utilsFile.get("temp_xls_file")
     
        content = [ ['音乐标示', '英文内容', '位置', '类型'], ['SoundName', 'Content', 'Pos', 'Body'] ]
        unitList = self.get_unit_list()
        for unit_folder in unitList: 
            if utils.is_meta(unit_folder):continue
            bXlsx.create_sheet(wb, unit_folder, content )

        bXlsx.delete_sheet(wb, "Sheet") # 删除默认的
        wb.save(temp_xls_file)

    
    def get_unit_list(self):
        arr = []
        book_unit_file = utilsFile.get("book_unit_file")
        for num, line in enumerate(open(book_unit_file, 'r',  encoding="utf-8")):
            if num < 2 : continue
            unit_name = line.split("\t")[1]
            arr.append(unit_name)

        return arr
    
    def reset_all(self):
        spk_en_audio_file = utilsFile.get("osd_en_audio_file")
        txt_folder_path = utilsFile.get("txt_folder_path")

        utils.mkdir(txt_folder_path)
        # filed = "音乐标示\t英文内容\t中文翻译\t位置\t行数\nSoundName\tContent\tChinese\tPos\tLine\n"
        # utils.create_text_file(spk_en_audio_file, filed)

        temp_audio_file = utilsFile.get("temp_audio_file")
        filed_word = "音乐标示\t英文内容\t中文翻译\t位置\t行数\nSoundName\tContent\tChinese\tPos\tLine\n"
        utils.create_text_file(temp_audio_file, filed_word)

    def xls2txt(self,  bTrans, restart,): 
        if restart: self.reset_all()       #如果没有中断,重置
        
        self.xlsx_path = utilsFile.get("temp_xls_file")  
        self.title_body_num = 1 
        self.end_body_num = 1
        self.end_body_idx = 0
    
        temp_audio_file = utilsFile.get("temp_audio_file")
        faudio = open(temp_audio_file, "a+",  encoding="utf-8")

        wb = openpyxl.load_workbook(self.xlsx_path)
        wss = wb.get_sheet_names()

        for sheetIdx in range(0, len(wss)):
            self.audioIdx = 1
            self.end_body_idx = 1
            self.cur_unit_name = ""
           
            self.unit_name = wss[sheetIdx]
            ws = wb.get_sheet_by_name(self.unit_name)  # 获取工作簿中的表格
            
            for idx in range(3, ws.max_row + 1):
                if idx <= self.end_body_idx: continue
                line = self.create_txt_line(ws, idx, bTrans)
                if line != "": faudio.write(line)
            
            if self.title_body_num != self.end_body_num : print("body num not match", self.title_body_num, self.end_body_num)
        faudio.close()



    def create_txt_line(self, ws, idx, bTrans):
        sound_file = ws.cell(row = idx, column = 1).value 
        english = ws.cell(row = idx, column = 2).value 
        pos = ws.cell(row = idx, column = 3).value
        vBody = ws.cell(row = idx, column = 4).value
        #body = int(ws.cell(row = idx, column = 4).value)

        body = 1
        #如果vbody的类型是int，那么就是数字，就用int函数转换
        if type(vBody) == int:
            body = int(vBody)
        else:
            body = 1

        
        if self.cur_unit_name != sound_file:
            self.cur_unit_name = sound_file
            self.audioIdx = 1
        else:
            self.audioIdx = self.audioIdx + 1

        if body == 0:
            self.title_body_num = self.title_body_num + 1
            english, pos, endIdx = self.merge_by_body(idx, english, pos)
            body = 1
            line = endIdx - idx + 1
        else:
            line = 1

        new_audio_name = sound_file + str(self.audioIdx)
        unit, page = utils.split_file_name(self.cur_unit_name)


        unit_type = contentMgr.get_unit_type(unit)
        if contentMgr.get_trans_type(unit_type) == "TRANS_SPLIT": #单词直接切割，不用翻译
            english, chinese = utilsWord.split_en_cn(english) 
        else :
            chinese =  "翻译文本"
            if bTrans: chinese = pTrans.trans(english)

        sLine = str(line)  # f"{line}"
        english = str(english)
        print("create_txt_line", new_audio_name, english, chinese, pos, line, sLine)
        print("create_txt_line", type(new_audio_name), type(english), type(chinese), type(pos), type(line), type(sLine))
        line = new_audio_name + "\t" + english + "\t" + chinese + "\t" + pos + "\t" + sLine + "\n"
        return line


    def merge_by_body(self, startIdx, start_english_txt, start_pos):
        endIdx = startIdx + 1
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb.get_sheet_by_name(self.unit_name)
        max_pos = start_pos

        for i in range(endIdx, ws.max_row + 1):

            english_txt = ws.cell(row = i, column = 2).value
            pos = ws.cell(row = i,column = 3).value
            next_body = int(ws.cell(row = i, column = 4).value) 
            start_english_txt = start_english_txt + " " + english_txt
            #print("合并文本", english_txt, pos,next_body,start_english_txt)
            max_pos = utils.cal_pos_max(max_pos, pos)
            
            if next_body == 2:
                self.end_body_num = self.end_body_num + 1
                endIdx = i
                break
        
        self.end_body_idx = endIdx
        return start_english_txt, max_pos, endIdx

############################### 整理上下文 #################################
    def resort(self):
        self.xlsx_path = utilsFile.get("temp_xls_file")  
        wb = openpyxl.load_workbook(self.xlsx_path)
        wss = wb.get_sheet_names()
        
        for sheetIdx in range(0, len(wss)):
            ws = wb.get_sheet_by_name(wss[sheetIdx]) 
            self.unit_name = ws.title
            unit_page_name = ws.cell(row = 3, column = 1).value
            print("split_file_name unit_type",self.xlsx_path, unit_page_name)            
            unit, page = utils.split_file_name(unit_page_name)
                
            unit_type = contentMgr.get_unit_type(unit)
            
            if contentMgr.get_sort_type(unit_type) == "TOP_BOTTOM": #排列方式
                self.resort_top_bottom(wb, ws)
            else:
                self.resort_left_right(wb, ws)
    
    def resort_top_bottom(self, wb, ws):
        self.split_page(ws)
        
        for i in range(3, ws.max_row + 1):
            unit_page_name = ws.cell(row = i, column = 1).value
            unit, page = utils.split_file_name(unit_page_name)
            print("resort_top_bottom ", unit_page_name, page)
            page_end_idx = self.get_page_end_idx(int(page[4:]))
            print("resort_top_bottom ", unit_page_name, i)
        
            #check_lesson_word()

            if i == int(page_end_idx) :continue
            self.swap_close_row(wb, ws, i, page_end_idx)
            

        path = utilsFile.get("temp_xls_file")
        wb.save(path)
        

 

    def split_page(self, ws):
        self.page_ends = []
        self.page_total_num = 0
        self.page_end_idx = self.find_page_end_idx(1, 3)
        
        for idx in range(self.page_end_idx, ws.max_row + 1):
            unit_page_name = ws.cell(row = idx, column = 1).value
            unit, page = utils.split_file_name(unit_page_name)
            
            if idx > self.page_end_idx:
                
                if self.page_total_num == 1:
                    self.page_ends.append(self.page_end_idx)
                
                self.page_total_num = 0
                self.page_end_idx = self.find_page_end_idx(page[4:], idx)
                self.page_ends.append(self.page_end_idx)
            
            self.page_total_num = self.page_total_num + 1

          
    def resort_left_right(self, wb, ws):
        self.sprit_left_and_right(ws)

        content = [ ['音乐标示', '英文内容', '位置', '类型'], ['SoundName', 'Content', 'Pos', 'Body'] ]
        bXlsx.create_sheet(wb, "temp", content )
        self.sheet_name = ws.title

        for page_idx in range(0, len(self.pagePos)):
            for block_idx in range(0, len(self.pagePos[page_idx])):
                for i in range(0, len(self.pagePos[page_idx][block_idx])):
                    
                    idx = self.pagePos[page_idx][block_idx][i]
                    value = bXlsx.get_row_value(ws.title, idx)
                    bXlsx.write_tag_xls_append(wb, "temp", value)
                    print("resort_left_right", value )

        path = utilsFile.get("temp_xls_file")
        wb.save(path)
        
        bXlsx.delete_sheet(wb, self.sheet_name)
        bXlsx.rename_sheet(wb, "temp", self.sheet_name)

    def sprit_left_and_right(self, ws):
        self.pagePos = []
       
        for i in range(3, ws.max_row + 1):
            unit_page_name = ws.cell(row = i, column = 1).value 
            unit, page = utils.split_file_name(unit_page_name)
            page_idx = int(page[4:]) - 1

            if len(self.pagePos) <= page_idx :
                self.pagePos.append([])
                #print(self.pagePos,unit,page, page_idx)
                self.pagePos[page_idx] = [[], [], []]

            pos = ws.cell(row = i, column = 3).value 
            eng_txt = ws.cell(row = i, column = 2).value 
            self.set_page_pos(page_idx, i, pos, eng_txt)

    def set_page_pos(self, page_idx, idx, pos, txt ):
        left = int(pos.split(",")[1])
        width = int(pos.split(",")[2])
        right = left + width
        mid = 530
        if left < mid and right < mid :     #left
            self.pagePos[page_idx][1].append(idx)
        elif left > mid and right > mid :   #right
            self.pagePos[page_idx][2].append(idx)
        else:                               #title            
            self.pagePos[page_idx][0].append(idx)       


    def find_page_end_idx(self, pageIdx, startIdx = 3):
        endIdx = startIdx
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb.get_sheet_by_name(self.unit_name)
        
        for i in range(startIdx + 1, ws.max_row + 1):
            unit_page_name = ws.cell(row = i, column = 1).value
            #print("find_page_end_idx", pageIdx, i, ws.max_row + 1, ws.title)
            unit, page = utils.split_file_name(unit_page_name)
            if int(page[4:]) > int(pageIdx): break
            endIdx = endIdx +1
        return endIdx
    
    def get_page_end_idx(self, page_idx):
  
        if self.page_ends[page_idx - 1] :
            return self.page_ends[page_idx - 1]
        '''
        else:
            print("get page end_idx failed",self.page_ends, page_idx - 1)
            sys.exit(1)
        '''

    #把相邻的行放到下面
    def swap_close_row(self, wb, ws, tag_idx, page_end_idx):
        #print("swap_close_row start", tag_idx, page_end_idx)
        for j in range(tag_idx + 1, page_end_idx + 1):
            #print("swap_close_row", tag_idx, j, self.is_close_context(tag_idx, j))
            if self.is_close_context(tag_idx, j):
                self.set_body_state(wb, ws, tag_idx, j)
                for k in range(j, tag_idx + 1, -1):
                    bXlsx.swap_row(wb, ws, k - 1, k)
                break

    def set_body_state(self, wb, ws, idx1, idx2):
        body1 = ws.cell(row = idx1, column = 4).value
        body2 = ws.cell(row = idx2, column = 4).value

        if (int(body1) == 1) & (int(body2) == 1):
            ws.cell(idx1, 4).value = 0
            ws.cell(idx2, 4).value = 2
        elif ((int(body1) == 2) & (int(body2) == 1)):
            ws.cell(idx1, 4).value = 1
            ws.cell(idx2, 4).value = 2

        wb.save(self.xlsx_path) 

    def is_close_context(self, idx1, idx2):
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb.get_sheet_by_name(self.unit_name)
    
        pos1 = ws.cell(row = idx1, column = 3).value
        pos2 = ws.cell(row = idx2, column = 3).value
        str1 = ws.cell(row = idx1, column = 2).value
        str2 = ws.cell(row = idx2, column = 2).value


        close_state = utils.is_pos_close(pos1, pos2)
        #font_state = utils.is_font_close(pos1, pos2, str1, str2)
        #if close_state and font_state: return True
        if close_state : return True
        return False

pXlsx = PelbsXlsx()