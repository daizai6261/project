import datetime
import os

from script.base.configer import configer
from script.utils.utils import utils
from script.utils.utilsfile import utilsFile

import openpyxl

class BaseXlsx:
    def __init__(self, key_word_excel_File, main_excel_file):
        self.key_word_excel_File = key_word_excel_File
        self.main_excel_file = main_excel_file

    def create_sheet(self, wb, sheet_name, init_content):
        ws = wb.create_sheet(sheet_name)   # 表命名
        for i in range(len(init_content)):
            for j in range(len(init_content[i])):
                ws.cell(row = i + 1, column = j +1).value = init_content[i][j]

        path = utilsFile.get("temp_xls_file") 
        wb.save(path)

    def delete_sheet(self, wb, sheet_name):
        path = utilsFile.get("temp_xls_file")
        ws = wb[sheet_name]
        wb.remove(ws)
        wb.save(path)
        

    def rename_sheet(self, wb, sheet_name, new_name):
        path = utilsFile.get("temp_xls_file")
        ws = wb[sheet_name]
        ws.title = new_name
        wb.save(path)


    # 获取某行所有值
    def get_row_value(self, sheet_name,  row):
        path = utilsFile.get("temp_xls_file")
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name]
        row_data = []
        for i in range(1, ws.max_column + 1):
            cell_value = ws.cell(row = row, column = i).value
            row_data.append(cell_value)
        return row_data
    
    def write_tag_xls_append(self, wb, sheet_name, value):
        #print("write_tag_xls_append ", value )
        path = utilsFile.get("temp_xls_file") 
        ws = wb.get_sheet_by_name(sheet_name)
        ws.append(value)
        wb.save(path)
    
    def append_xls_value(self, sheet_name, value):
        
        path = utilsFile.get("temp_xls_file")
        wb = openpyxl.load_workbook(path) 
        ws = wb.get_sheet_by_name(sheet_name)
        ws.append(value)
        wb.save(path)

    def swap_row(self, wb , ws, idx1, idx2):
        for i in range(1, ws.max_column + 1):
            value1 = ws.cell(row = idx1, column = i).value
            value2 = ws.cell(row = idx2, column = i).value
            ws.cell(row = idx1, column = i).value = value2
            ws.cell(row = idx2, column = i).value = value1
        path = utilsFile.get("temp_xls_file")
        wb.save(path)

bXlsx = BaseXlsx("1", 1)